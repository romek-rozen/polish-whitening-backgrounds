"""Merge the two clustering sweeps into one pivot-ready Excel workbook.

Both families answer the same question — how many groups, and how much of the
list is noise — but they are **not parameter variants of each other**, so the
workbook keeps a ``method`` column rather than pretending they are comparable
row-for-row:

- ``knn_leiden``          — kNN graph + Leiden + similarity floor + min size
- ``threshold_unionfind`` — the keyword_cluster tier: cosine threshold,
                            connected components, min cluster size

Sheets:

- **sweep** — every run, one row each. This is the pivot source.
- **legend** — what each column means and the two traps in reading it.

Usage::

    python make_xlsx.py --leiden work/noise_sweep.csv \\
        --threshold work/threshold_sweep.csv --out work/klastrowanie.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

COLUMNS = [
    "method", "model", "space", "floor_mode", "floor_setting", "floor_cosine",
    "resolution", "min_size", "n_keywords", "n_clusters", "n_noise", "noise_pct",
    "largest", "median_size", "clusters_lt_10", "mean_edge_sim",
    "edges_kept_pct", "isolated_nodes", "n_edges", "components_before_min_size",
    "umap_dims", "umap_neighbors",
]

LEGEND = [
    ("method", "knn_leiden = kNN graph + Leiden. threshold_unionfind = keyword_cluster: "
               "cosine threshold + connected components. umap_hdbscan = UMAP projection + "
               "density clustering. Three different families, not settings of one."),
    ("space", "raw = plain L2-normalised embeddings. background = after applying the "
              "fitted ZCA whitening background for that model."),
    ("floor_mode", "abs = absolute cosine cutoff. pct = drop the weakest X% of edges."),
    ("floor_setting", "The knob value: cosine for abs, percentage for pct."),
    ("floor_cosine", "The cosine actually used as the cutoff."),
    ("resolution", "Leiden resolution. Higher = more, smaller groups. Empty for the "
                   "threshold method, which has no such knob."),
    ("min_size", "Clusters smaller than this are relabelled as noise."),
    ("n_clusters", "Groups produced, excluding noise."),
    ("n_noise", "Keywords assigned to no group."),
    ("noise_pct", "n_noise as a percentage of all keywords."),
    ("largest", "Size of the biggest group. A huge value means one bucket swallowed "
                "the list — check it before trusting the rest."),
    ("median_size", "Median group size."),
    ("clusters_lt_10", "Groups with fewer than 10 keywords."),
    ("mean_edge_sim", "Mean cosine over kNN edges in that space (knn_leiden only)."),
    ("edges_kept_pct", "Share of kNN edges surviving the floor (knn_leiden only)."),
    ("isolated_nodes", "Keywords left with no edge after the floor (knn_leiden only)."),
    ("n_edges", "Pairs above threshold (threshold_unionfind only)."),
    ("components_before_min_size", "Connected components before the min-size filter "
                                   "(threshold_unionfind only)."),
    ("", ""),
    ("TRAP 1", "Do NOT compare abs floors across spaces. Whitening rescales every "
               "cosine, so the same number means something different in raw vs "
               "background. Use the pct rows for cross-space comparison."),
    ("TRAP 2", "Noise is not a defect to minimise. floor=0 + min_size=1 gives 0% noise "
               "because every keyword is forced into a group, including junk that "
               "belongs nowhere. Some noise means the method is being honest."),
]


def read_rows(path: Path, method: str | None) -> list[dict]:
    if not path.exists():
        print(f"  WARNING: {path} missing — skipped")
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row.setdefault("method", method or "")
        if not row.get("method"):
            row["method"] = method or ""
    print(f"  {path.name}: {len(rows)} rows")
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--leiden", type=Path, required=True)
    ap.add_argument("--threshold", type=Path, required=True)
    ap.add_argument("--umap", type=Path, default=None)
    ap.add_argument("--out", type=Path, required=True)
    args = ap.parse_args()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    print("reading:")
    rows = read_rows(args.leiden, "knn_leiden") + read_rows(args.threshold, "threshold_unionfind")
    if args.umap:
        rows += read_rows(args.umap, "umap_hdbscan")
    if not rows:
        raise SystemExit("no rows to write")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sweep"
    sheet.append(COLUMNS)

    numeric = {
        "floor_setting", "floor_cosine", "resolution", "min_size", "n_keywords",
        "n_clusters", "n_noise", "noise_pct", "largest", "median_size",
        "clusters_lt_10", "mean_edge_sim", "edges_kept_pct", "isolated_nodes",
        "n_edges", "components_before_min_size", "umap_dims", "umap_neighbors",
    }
    for row in rows:
        out = []
        for column in COLUMNS:
            value = row.get(column, "")
            if column in numeric and value not in ("", None):
                try:
                    value = float(value)
                    if value.is_integer():
                        value = int(value)
                except (TypeError, ValueError):
                    pass
            out.append(value)
        sheet.append(out)

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(11, len(column) + 2)

    legend = workbook.create_sheet("legend")
    legend.append(["column", "meaning"])
    for cell in legend[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for name, meaning in LEGEND:
        legend.append([name, meaning])
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 110
    for row_cells in legend.iter_rows(min_row=2):
        row_cells[1].alignment = Alignment(wrap_text=True, vertical="top")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"\nwrote {args.out}  ({len(rows)} rows, {len(COLUMNS)} columns)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
