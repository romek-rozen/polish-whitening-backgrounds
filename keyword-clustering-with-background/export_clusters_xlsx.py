"""Run every clustering method once and export the actual clusters to Excel.

The sweeps answer "how many groups and how much noise". They say nothing about
whether the groups are usable, which is the only question that matters when the
output becomes ad groups. This produces something you can read:

- **keywords** — one row per keyword, one column pair per method, so you can
  filter a keyword and see where each method put it.
- **clusters** — one row per cluster: method, label, size, and a sample of
  members. This is the sheet to skim.
- **summary** — cluster counts and noise per method.

Cluster labels are the keyword closest to the centroid — the least idiosyncratic
member, which reads as the group's theme.

Uses cached embeddings, so re-running with different settings is cheap.

Usage::

    python export_clusters_xlsx.py --csv work/keywords.csv \\
        --model bge_m3 --out work/klastry_podglad.xlsx
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent / "pl-keyword-embedding-cpu-bench"))

from bench import MODELS, apply_background, l2_normalize  # noqa: E402
from cluster_keywords import cluster as leiden_cluster  # noqa: E402
from cluster_keywords import cluster_labels, load_keywords  # noqa: E402
from embed_cache import encode_cached  # noqa: E402
from sweep_leiden import BACKGROUNDS, apply_min_size, knn_edges, leiden  # noqa: E402
from sweep_threshold import apply_min_cluster_size, threshold_components  # noqa: E402
from sweep_umap_hdbscan import hdbscan_labels, umap_reduce  # noqa: E402
from cluster_agglomerative import agglomerative, demote_small  # noqa: E402


def run_methods(vectors: np.ndarray, seed: int) -> dict[str, np.ndarray]:
    """One representative configuration per family, chosen from the sweeps."""
    out: dict[str, np.ndarray] = {}

    # Leiden without a cutoff: labels everything, 0% noise by construction.
    out["leiden_r4"] = leiden_cluster(vectors, 15, 4.0, seed)

    # Leiden with the sharpening D91 uses: weakest 25% of edges dropped,
    # clusters under 10 keywords demoted to noise.
    lo, hi, sim = knn_edges(vectors, 15, seed, 4)
    cut = float(np.percentile(sim, 25.0))
    keep = sim >= cut
    labels = leiden(len(vectors), lo[keep], hi[keep], sim[keep], 4.0, seed)
    connected = np.zeros(len(vectors), dtype=bool)
    connected[lo[keep]] = True
    connected[hi[keep]] = True
    out["leiden_sharpened"] = apply_min_size(np.where(connected, labels, -1), 10)

    # Density clustering on a UMAP projection: noise is decided, not configured.
    points = umap_reduce(vectors, 30, 15, seed)
    out["umap_hdbscan"] = hdbscan_labels(points, 10)

    # The keyword_cluster semantic tier at its documented default. Union-find
    # over a threshold is single linkage cut at that threshold.
    components, _ = threshold_components(vectors, 0.80)
    out["bdos_threshold_0.8"] = apply_min_cluster_size(components, 2)

    # Same idea, average linkage: merge closest first, stop at 0.5. Average
    # linkage is what stops one bridging keyword from chaining two groups.
    out["agglomerative_0.5"] = demote_small(
        agglomerative(vectors, 0.5, "average"), 2)
    return out


def cluster_rows(keywords: list[str], vectors: np.ndarray, method: str,
                 labels: np.ndarray, sample: int) -> list[dict]:
    members: dict[int, list[int]] = defaultdict(list)
    for row, cluster_id in enumerate(labels):
        members[int(cluster_id)].append(row)
    noise = members.pop(-1, [])

    names = cluster_labels(keywords, vectors, labels)
    rows = []
    for cluster_id in sorted(members, key=lambda c: -len(members[c])):
        rows_in = members[cluster_id]
        rows.append({
            "method": method,
            "cluster": cluster_id,
            "label": names.get(cluster_id, ""),
            "size": len(rows_in),
            "sample": " | ".join(keywords[r] for r in rows_in[:sample]),
        })
    if noise:
        rows.append({
            "method": method, "cluster": -1, "label": "(NOISE)",
            "size": len(noise),
            "sample": " | ".join(keywords[r] for r in noise[:sample]),
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--csv", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--models", nargs="*", default=list(MODELS), choices=list(MODELS))
    ap.add_argument("--backgrounds", type=Path, default=ROOT.parent / "backgrounds")
    ap.add_argument("--no-background", action="store_true")
    ap.add_argument("--device", default="cuda")
    ap.add_argument("--batch-size", type=int, default=64)
    ap.add_argument("--refresh-cache", action="store_true")
    ap.add_argument("--sample", type=int, default=15, help="members shown per cluster")
    ap.add_argument("--dump-dir", type=Path, default=None,
                    help="Also write one plain-text dump per method (for review).")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    import logging
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    keywords = load_keywords(args.csv)
    print(f"{len(keywords)} unique keywords")

    # (model, method) -> labels, plus the vectors each was clustered in, so
    # cluster labels are computed in the same space that produced them.
    results: dict[tuple[str, str], np.ndarray] = {}
    spaces: dict[str, np.ndarray] = {}
    for model_key in args.models:
        encoded = encode_cached(model_key, keywords, batch_size=args.batch_size,
                                device=args.device, refresh=args.refresh_cache)
        if args.no_background:
            vectors, space = l2_normalize(encoded.vectors), "raw"
        else:
            bg = args.backgrounds / BACKGROUNDS[model_key]
            vectors, space = apply_background(encoded.vectors, bg), bg.name
        print(f"{model_key}: space={space}")
        spaces[model_key] = vectors
        for method, labels in run_methods(vectors, args.seed).items():
            results[(model_key, method)] = labels
        print()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    head_fill = PatternFill("solid", fgColor="DDDDDD")

    # --- clusters sheet: the one to skim -----------------------------------
    sheet = workbook.active
    sheet.title = "clusters"
    sheet.append(["model", "method", "cluster", "label", "size", "sample_keywords"])
    all_cluster_rows = []
    for (model_key, method), labels in results.items():
        rows = cluster_rows(keywords, spaces[model_key], method, labels, args.sample)
        for row in rows:
            row["model"] = model_key
            all_cluster_rows.append(row)
            sheet.append([model_key, row["method"], row["cluster"], row["label"],
                          row["size"], row["sample"]])
    for width, letter in zip((16, 22, 9, 34, 8, 160), "ABCDEF"):
        sheet.column_dimensions[letter].width = width

    # --- keywords sheet: per-keyword comparison ----------------------------
    kw_sheet = workbook.create_sheet("keywords")
    header = ["keyword"]
    for model_key, method in results:
        header.append(f"{model_key}__{method}")
    kw_sheet.append(header)

    names = {k: cluster_labels(keywords, spaces[k[0]], l) for k, l in results.items()}
    sizes = {k: Counter(l.tolist()) for k, l in results.items()}
    for i, keyword in enumerate(keywords):
        row = [keyword]
        for key, labels in results.items():
            cid = int(labels[i])
            row.append("(NOISE)" if cid == -1 else names[key].get(cid, ""))
        kw_sheet.append(row)
    kw_sheet.column_dimensions["A"].width = 42
    for index in range(2, len(header) + 1):
        kw_sheet.column_dimensions[get_column_letter(index)].width = 30

    # --- summary sheet -----------------------------------------------------
    summary = workbook.create_sheet("summary")
    summary.append(["model", "method", "clusters", "noise", "noise_%", "largest", "median_size"])
    for (model_key, method), labels in results.items():
        real = labels[labels >= 0]
        counts = np.bincount(real) if real.size else np.array([0])
        counts = counts[counts > 0]
        noise = int((labels == -1).sum())
        summary.append([
            model_key, method, int(counts.size), noise,
            round(100 * noise / len(keywords), 2),
            int(counts.max()) if counts.size else 0,
            int(np.median(counts)) if counts.size else 0,
        ])
    for letter, width in zip("ABCDEFG", (16, 22, 10, 10, 10, 10, 13)):
        summary.column_dimensions[letter].width = width

    for page in (sheet, kw_sheet, summary):
        for cell in page[1]:
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")
        page.freeze_panes = "A2"
        page.auto_filter.ref = page.dimensions

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"wrote {args.out}")

    if args.dump_dir:
        args.dump_dir.mkdir(parents=True, exist_ok=True)
        for model_key, method in results:
            rows = [r for r in all_cluster_rows
                    if r["method"] == method and r["model"] == model_key]
            lines = [f"# {model_key} / {method}   ({len(rows)} clusters incl. noise)", ""]
            for row in rows:
                lines.append(f"[{row['cluster']}] n={row['size']}  ~{row['label']}")
                for keyword in row["sample"].split(" | "):
                    lines.append(f"    {keyword}")
                lines.append("")
            path = args.dump_dir / f"{model_key}__{method}.txt"
            path.write_text("\n".join(lines), encoding="utf-8")
            print(f"  dump: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
