"""Produce the final production clustering workbook and import CSV."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent / "pl-keyword-embedding-cpu-bench"))

from bench import apply_background  # noqa: E402
from cluster_agglomerative import (  # noqa: E402
    agglomerative_tree,
    cut_agglomerative_tree,
    demote_small,
)
from cluster_keywords import cluster_labels, load_keywords  # noqa: E402
from embed_cache import cache_path, corpus_fingerprint, encode_cached  # noqa: E402

MODEL = "bge_m3"
BACKGROUND = "bgem3_pl_kwmix900k_mrl1024"
LINKAGE = "average"
THRESHOLD = 0.526
MIN_SIZE = 2


def style_sheet(sheet, widths: tuple[float, ...]) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def build_workbook(path: Path, keywords: list[str], labels: np.ndarray,
                   vectors: np.ndarray) -> None:
    names = cluster_labels(keywords, vectors, labels)
    sizes = Counter(int(label) for label in labels)
    members: dict[int, list[str]] = defaultdict(list)
    for keyword, label in zip(keywords, labels):
        members[int(label)].append(keyword)

    workbook = Workbook()
    groups = workbook.active
    groups.title = "grupy"
    groups.append(["cluster_id", "label", "size", "keywords"])
    cluster_ids = sorted(
        (cluster_id for cluster_id in members if cluster_id != -1),
        key=lambda cluster_id: (-sizes[cluster_id], names[cluster_id], cluster_id),
    )
    for cluster_id in cluster_ids:
        groups.append([
            cluster_id,
            names[cluster_id],
            sizes[cluster_id],
            " | ".join(sorted(members[cluster_id])),
        ])
    style_sheet(groups, (13, 38, 10, 150))
    groups.column_dimensions["D"].width = 150
    for cell in groups["D"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    phrases = workbook.create_sheet("frazy")
    phrases.append(["keyword", "cluster_id", "label", "size", "is_noise"])
    for keyword, label in zip(keywords, labels):
        cluster_id = int(label)
        is_noise = cluster_id == -1
        phrases.append([
            keyword,
            cluster_id,
            "(NOISE)" if is_noise else names[cluster_id],
            sizes[cluster_id],
            is_noise,
        ])
    style_sheet(phrases, (45, 13, 40, 10, 12))

    real_sizes = np.asarray(
        [size for cluster_id, size in sizes.items() if cluster_id != -1], dtype=int
    )
    noise = sizes.get(-1, 0)
    info = workbook.create_sheet("info")
    info.append(["parameter", "value"])
    info_rows = [
        ("generated_at", datetime.now(ZoneInfo("Europe/Warsaw")).isoformat(timespec="seconds")),
        ("input", "work/keywords_url_2026-07-23.csv"),
        ("model", MODEL),
        ("background", BACKGROUND),
        ("space", f"{MODEL} + {BACKGROUND}"),
        ("algorithm", "agglomerative clustering"),
        ("linkage", LINKAGE),
        ("similarity_threshold", THRESHOLD),
        ("min_size", MIN_SIZE),
        ("keywords", len(keywords)),
        ("clusters", len(real_sizes)),
        ("noise", noise),
        ("noise_pct", round(100 * noise / len(keywords), 2)),
        ("largest_cluster", int(real_sizes.max()) if len(real_sizes) else 0),
        ("median_cluster_size", float(np.median(real_sizes)) if len(real_sizes) else 0),
        (
            "noise_warning",
            "Szum (~20%) to NIE są śmieci: to frazy bez bliskiego bliźniaka; są tam "
            "frazy ofertowe. NIE kasować bez przejrzenia.",
        ),
        (
            "pre_export_review",
            "Przed użyciem jako grupy reklam warto odsiać nazwiska i modyfikatory "
            "geograficzne. Wszystkie modele robią z nazwisk jeden worek, a geo przeciąga "
            "frazę do klastra lokalnego.",
        ),
        (
            "threshold_scope",
            "Próg 0.526 dotyczy WYŁĄCZNIE tej przestrzeni: bge-m3 + background "
            "bgem3_pl_kwmix900k_mrl1024. Nie przenosi się na surową przestrzeń ani "
            "na inny model.",
        ),
    ]
    for row in info_rows:
        info.append(row)
    style_sheet(info, (25, 120))
    for cell in info["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_csv(path: Path, keywords: list[str], labels: np.ndarray,
              names: dict[int, str]) -> None:
    sizes = Counter(int(label) for label in labels)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["keyword", "cluster_id", "label", "size"])
        for keyword, label in zip(keywords, labels):
            cluster_id = int(label)
            writer.writerow([
                keyword,
                cluster_id,
                "(NOISE)" if cluster_id == -1 else names[cluster_id],
                sizes[cluster_id],
            ])


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path,
                        default=ROOT / "work/keywords_url_2026-07-23.csv")
    parser.add_argument("--out-xlsx", type=Path,
                        default=ROOT / "work/klastry_FINAL.xlsx")
    parser.add_argument("--out-csv", type=Path,
                        default=ROOT / "work/klastry_FINAL.csv")
    parser.add_argument("--backgrounds", type=Path,
                        default=ROOT.parent / "backgrounds")
    parser.add_argument("--cache-dir", type=Path,
                        default=ROOT / "work/embeddings")
    args = parser.parse_args()

    keywords = load_keywords(args.csv)
    fingerprint = corpus_fingerprint(keywords)
    expected = cache_path(args.cache_dir, MODEL, fingerprint)
    if not expected.exists():
        raise SystemExit(f"Missing embedding cache; refusing to re-embed: {expected}")
    encoded = encode_cached(MODEL, keywords, device="cpu", cache_dir=args.cache_dir)
    vectors = apply_background(encoded.vectors, args.backgrounds / BACKGROUND)
    tree = agglomerative_tree(vectors, LINKAGE)
    labels = demote_small(cut_agglomerative_tree(tree, THRESHOLD), MIN_SIZE)
    names = cluster_labels(keywords, vectors, labels)

    build_workbook(args.out_xlsx, keywords, labels, vectors)
    write_csv(args.out_csv, keywords, labels, names)
    print(f"wrote {args.out_xlsx}")
    print(f"wrote {args.out_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
